"""استخراج المصطلحات — بدائل عن الإدخال كلمة كلمة.

تلات مصادر، مرتّبة من الأرخص للأغلى:

1. **من ذاكرة الترجمة** — الأزواج المعتمدة عندك بالفعل. مفيش رفع ولا
   انتظار، والبيانات موجودة أصلًا. ده أول مكان تدوّر فيه.
2. **استيراد جدول** — CSV أو Excel بعمودين. مفيش أي نداء API.
3. **من ملف وأصله المترجَم** — بنحاذي الملفين بالمقاطع وبنستخرج
   الأزواج. الأقوى لو عندك أرشيف ترجمات قديمة.

مبدأ حاكم: **المستخرَج ماينضمّش للقاعدة تلقائيًا**. بيرجع كمرشّحين
والمستخدم يوافق. نفس منطق الانتشار: المصطلح الغلط في القاعدة بيتحقن
في تعليمات كل ترجمة بعد كده، فتكلفة الخطأ أعلى من تكلفة المراجعة.
"""
from __future__ import annotations

import csv
import io
import json
import logging
import re
from collections import Counter
from dataclasses import dataclass, field
from pathlib import Path

from app.core.config import settings
from app.tools.translator.costing import compute_cost
from app.tools.translator.engine import Usage
from app.tools.translator.formats.base import strip_tags
from app.tools.translator.langs import language_name

logger = logging.getLogger(__name__)


@dataclass
class TermCandidate:
    source_term: str
    target_term: str
    # كام مرة ظهر الزوج ده في المادة المفحوصة
    frequency: int = 1
    # جملة نموذجية بيظهر فيها — بتساعد المراجع يقرر
    sample: str = ""
    note: str = ""


@dataclass
class ExtractionResult:
    candidates: list[TermCandidate] = field(default_factory=list)
    pairs_examined: int = 0
    usage: Usage = field(default_factory=Usage)
    warnings: list[str] = field(default_factory=list)


# ---------------------------------------------------------------------------
# 1) استيراد جدول (CSV / Excel) — بدون أي نداء API
# ---------------------------------------------------------------------------
# أسماء أعمدة شائعة، عشان نتعرّف على الجدول من غير ما نسأل المستخدم
_SOURCE_HEADERS = {
    "source", "term", "arabic", "ar", "source_term", "المصطلح", "المصدر",
    "عربي", "العربية", "kaynak", "источник", "термин",
}
_TARGET_HEADERS = {
    "target", "translation", "english", "en", "target_term", "الترجمة",
    "الهدف", "إنجليزي", "الإنجليزية", "hedef", "перевод",
}


def _looks_like_header(first_row: list[str]) -> bool:
    lowered = {str(cell).strip().lower() for cell in first_row if cell}
    return bool(lowered & _SOURCE_HEADERS or lowered & _TARGET_HEADERS)


def _rows_from_csv(data: bytes) -> list[list[str]]:
    # الملفات العربية القديمة كتير بتتحفظ بـ cp1256
    for encoding in ("utf-8-sig", "utf-8", "cp1256", "latin-1"):
        try:
            text = data.decode(encoding)
            break
        except UnicodeDecodeError:
            continue
    else:
        text = data.decode("utf-8", errors="replace")

    sample = text[:4096]
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=",;\t")
    except csv.Error:
        dialect = csv.excel

    return [row for row in csv.reader(io.StringIO(text), dialect) if any(row)]


def _rows_from_xlsx(path: Path) -> list[list[str]]:
    from openpyxl import load_workbook

    workbook = load_workbook(str(path), read_only=True, data_only=True)
    rows: list[list[str]] = []
    for sheet in workbook.worksheets:
        for values in sheet.iter_rows(values_only=True):
            cells = ["" if v is None else str(v).strip() for v in values]
            if any(cells):
                rows.append(cells)
    workbook.close()
    return rows


def import_table(path: Path, max_terms: int = 5000) -> ExtractionResult:
    """قراءة مصطلحات من جدول بعمودين.

    بنقبل CSV و Excel. أول صف بيتتخطّى لو شكله عناوين أعمدة.
    """
    result = ExtractionResult()
    suffix = path.suffix.lower()

    if suffix in (".xlsx", ".xlsm"):
        rows = _rows_from_xlsx(path)
    elif suffix in (".csv", ".txt", ".tsv"):
        rows = _rows_from_csv(path.read_bytes())
    else:
        result.warnings.append(f"صيغة غير مدعومة للاستيراد: {suffix}")
        return result

    if not rows:
        result.warnings.append("الملف فاضي")
        return result

    if _looks_like_header(rows[0]):
        rows = rows[1:]

    seen: set[tuple[str, str]] = set()
    for row in rows:
        if len(row) < 2:
            continue
        source = str(row[0]).strip()
        target = str(row[1]).strip()
        note = str(row[2]).strip() if len(row) > 2 else ""

        if not source or not target:
            continue
        # الصفوف الطويلة جمل مش مصطلحات
        if len(source) > 120 or len(target) > 120:
            continue
        if (source, target) in seen:
            continue
        seen.add((source, target))

        result.candidates.append(
            TermCandidate(source_term=source, target_term=target, note=note)
        )
        if len(result.candidates) >= max_terms:
            result.warnings.append(f"وقفنا عند {max_terms} مصطلح")
            break

    result.pairs_examined = len(rows)
    return result


# ---------------------------------------------------------------------------
# 2) الاستخراج من أزواج مترجمة (ذاكرة أو ملفين)
# ---------------------------------------------------------------------------
_SCHEMA = {
    "type": "object",
    "properties": {
        "terms": {
            "type": "array",
            "items": {
                "type": "object",
                "properties": {
                    "source_term": {"type": "string"},
                    "target_term": {"type": "string"},
                    "why": {"type": "string"},
                },
                "required": ["source_term", "target_term", "why"],
                "additionalProperties": False,
            },
        }
    },
    "required": ["terms"],
    "additionalProperties": False,
}

_SYSTEM = """\
You build terminology glossaries from translated text.

You are given aligned sentence pairs. Extract the **terms** — the words and
short phrases whose rendering should stay fixed across a document — and give
each one's established rendering as it appears in the pairs.

Include:
- Domain terminology (legal, medical, technical, scientific).
- Defined parties, roles, offices, and institution names.
- Product names, document types, and form labels that recur.
- Any phrase whose translation is not the obvious literal one.

Exclude:
- Whole sentences or clauses. A term is normally one to four words.
- Ordinary vocabulary a competent translator would never vary.
- Anything appearing only once and carrying no domain weight.
- Numbers, dates, and codes.

For each term give the base form, not an inflected one, and give the
rendering exactly as the translator used it. `why` is a short reason the
term belongs in a glossary — one clause, no more.

Return nothing if the pairs contain no terminology worth fixing. An empty
glossary is better than a padded one: every entry is injected into the
instructions for every later translation.
"""


def _build_prompt(
    pairs: list[tuple[str, str]], source_lang: str, target_lang: str
) -> str:
    lines = [
        f"Source language: {language_name(source_lang)}",
        f"Target language: {language_name(target_lang)}",
        "",
        "Aligned pairs:",
    ]
    for index, (source, target) in enumerate(pairs, start=1):
        lines.append(f"{index}. {source}")
        lines.append(f"   -> {target}")
    return "\n".join(lines)


def extract_from_pairs(
    pairs: list[tuple[str, str]],
    source_lang: str = "ar",
    target_lang: str = "en",
    domain: str = "general",
    model: str | None = None,
    batch_size: int = 40,
    api_key: str | None = None,
) -> ExtractionResult:
    """استخراج مصطلحات من أزواج جمل مترجمة."""
    result = ExtractionResult(pairs_examined=len(pairs))
    if not pairs:
        return result

    key = api_key or settings.anthropic_api_key
    if not key:
        result.warnings.append(
            "استخراج المصطلحات محتاج مفتاح Anthropic. "
            "الاستيراد من جدول شغّال من غيره."
        )
        return result

    import anthropic

    client = anthropic.Anthropic(api_key=key)
    model = model or settings.default_model

    system = _SYSTEM
    if domain != "general":
        system += f"\nThe documents are in the {domain} domain.\n"

    # نجمّع النتائج عبر الدفعات ونحسب التكرار — المصطلح اللي بيتكرر
    # في دفعات مختلفة أجدر بالثقة
    tally: Counter[tuple[str, str]] = Counter()
    notes: dict[tuple[str, str], str] = {}
    samples: dict[tuple[str, str], str] = {}

    for start in range(0, len(pairs), batch_size):
        chunk = pairs[start : start + batch_size]
        try:
            response = client.messages.create(
                model=model,
                max_tokens=4000,
                system=[
                    {
                        "type": "text",
                        "text": system,
                        "cache_control": {"type": "ephemeral"},
                    }
                ],
                thinking={"type": "adaptive"},
                output_config={
                    "effort": "medium",
                    "format": {"type": "json_schema", "schema": _SCHEMA},
                },
                messages=[
                    {
                        "role": "user",
                        "content": _build_prompt(chunk, source_lang, target_lang),
                    }
                ],
            )
        except Exception as exc:  # noqa: BLE001
            logger.exception("فشل استخراج المصطلحات")
            result.warnings.append(f"{type(exc).__name__}: {exc}")
            continue

        raw = response.usage
        usage = Usage(
            input_tokens=getattr(raw, "input_tokens", 0) or 0,
            output_tokens=getattr(raw, "output_tokens", 0) or 0,
            cache_read_tokens=getattr(raw, "cache_read_input_tokens", 0) or 0,
            cache_write_tokens=getattr(raw, "cache_creation_input_tokens", 0) or 0,
            calls=1,
        )
        usage.cost_usd = compute_cost(
            model,
            input_tokens=usage.input_tokens,
            output_tokens=usage.output_tokens,
            cache_read_tokens=usage.cache_read_tokens,
            cache_write_tokens=usage.cache_write_tokens,
        )
        result.usage.add(usage)

        text = next(
            (block.text for block in response.content if block.type == "text"), ""
        )
        if not text.strip():
            continue
        try:
            parsed = json.loads(text)
        except json.JSONDecodeError:
            result.warnings.append("رد غير قابل للقراءة من النموذج")
            continue

        for item in parsed.get("terms", []):
            source_term = str(item.get("source_term", "")).strip()
            target_term = str(item.get("target_term", "")).strip()
            if not source_term or not target_term:
                continue
            pair_key = (source_term, target_term)
            tally[pair_key] += 1
            notes.setdefault(pair_key, str(item.get("why", "")).strip())
            if pair_key not in samples:
                for source, target in chunk:
                    if source_term in source:
                        samples[pair_key] = source[:160]
                        break

    for (source_term, target_term), count in tally.most_common():
        result.candidates.append(
            TermCandidate(
                source_term=source_term,
                target_term=target_term,
                frequency=count,
                sample=samples.get((source_term, target_term), ""),
                note=notes.get((source_term, target_term), ""),
            )
        )

    return result


# ---------------------------------------------------------------------------
# 3) محاذاة ملف بترجمته
# ---------------------------------------------------------------------------
@dataclass
class Alignment:
    pairs: list[tuple[str, str]] = field(default_factory=list)
    source_units: int = 0
    target_units: int = 0
    confident: bool = True
    warnings: list[str] = field(default_factory=list)


def align_documents(
    source_path: Path, target_path: Path, work_dir: Path
) -> Alignment:
    """محاذاة مستند بترجمته عن طريق ترتيب الوحدات.

    الافتراض: الترجمة حافظت على بنية المستند (فقرة مقابل فقرة). ده
    صحيح في أغلب الترجمات المهنية لأن المترجم بيشتغل على نسخة من
    الملف الأصلي.

    لو عدد الوحدات مختلف بشكل كبير، بنعلّم النتيجة كغير موثوقة بدل ما
    نطلع أزواجًا متزحلقة — زوج غلط واحد بيولّد مصطلحًا غلط في القاعدة.
    """
    from app.tools.translator.formats import registry

    alignment = Alignment()
    work_dir.mkdir(parents=True, exist_ok=True)

    def units_of(path: Path) -> list[str]:
        fmt = registry.detect_format(path.name)
        working, working_fmt, meta = registry.prepare(fmt, path, work_dir)
        if meta.get("needs_ocr"):
            raise RuntimeError(
                f"«{path.name}» ممسوح ضوئيًا — شغّله كمشروع عادي الأول "
                "عشان يتقرا، وبعدين استخرج منه."
            )
        extracted = registry.extract(
            working_fmt, Path(working), normalize=(fmt == "pdf")
        )
        return [
            strip_tags(unit.text).strip()
            for unit in extracted.units
            if strip_tags(unit.text).strip()
        ]

    source_units = units_of(source_path)
    target_units = units_of(target_path)
    alignment.source_units = len(source_units)
    alignment.target_units = len(target_units)

    if not source_units or not target_units:
        alignment.warnings.append("واحد من الملفين مافيهوش نص")
        alignment.confident = False
        return alignment

    ratio = min(len(source_units), len(target_units)) / max(
        len(source_units), len(target_units)
    )
    if ratio < 0.85:
        alignment.confident = False
        alignment.warnings.append(
            f"عدد الفقرات مختلف كتير ({len(source_units)} مقابل "
            f"{len(target_units)}) — المحاذاة بالترتيب مش موثوقة هنا. "
            "المصطلحات المستخرجة محتاجة مراجعة أدق."
        )

    for source, target in zip(source_units, target_units):
        # الفقرات القصيرة جدًا (أرقام، رؤوس جداول) مش بتفيد الاستخراج
        if len(source) < 8 or len(target) < 8:
            continue
        alignment.pairs.append((source, target))

    return alignment
